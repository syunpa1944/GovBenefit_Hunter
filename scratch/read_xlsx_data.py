import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\럭포마_개발자료\앱인토스\공공데이터api연결\공공데이터api_활용신청_상세기능정보.xlsx"
print("File exists:", os.path.exists(file_path))

try:
    wb = openpyxl.load_workbook(file_path)
    print("Sheets in workbook:", wb.sheetnames)
    
    for name in wb.sheetnames:
        print(f"\n--- Sheet: {name} ---")
        sheet = wb[name]
        # print first 20 rows
        for r in range(1, 40):
            row_vals = [sheet.cell(r, c).value for c in range(1, 15)]
            # filter out completely empty rows
            if any(val is not None for val in row_vals):
                print(f"Row {r}:", row_vals)
except Exception as e:
    print("Error reading excel:", e)
